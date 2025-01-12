import pandas as pd
import numpy as np
import re
from openpyxl import Workbook
from matplotlib import pyplot as plt
from collections import Counter

def prof_year_state_prof(data, profession):
    years = {}
    for index in data:
        if profession and (
                ' ios' in index[0].lower() or
                'ios ' in index[0].lower() or
                '-ios' in index[0].lower() or
                'ios-' in index[0].lower() or
                'ios/' in index[0].lower() or
                '/ios' in index[0].lower()):
            year = index[6].year

            if year not in years:
                years[year] = [0, 0]
            
            salary_rub = index[7]

            if salary_rub > 0:
                years[year][0] += salary_rub
                years[year][1] += 1

    return [
        [key, round(value[0] / value[1]) if value[1] > 0 else 0, value[1]]
        for key, value in dict(sorted(years.items())).items()
    ]


def prof_year_state_all(data):
    years = {}
    for index in data:
        year = index[6].year
        
        if year not in years:
            years[year] = [0, 0]

        salary_rub = index[7]

        if salary_rub > 0 and salary_rub <= 10_000_000:
            years[year][0] += salary_rub
            years[year][1] += 1

    return [
        [key, round(value[0] / value[1]) if value[1] > 0 else 0, value[1]]
        for key, value in dict(sorted(years.items())).items()
    ]


def vacancy_count_by_year(data, profession=None):
    years = Counter()
    for index in data:
        if profession:
            if (' ios' in index[0].lower() or
                'ios ' in index[0].lower() or
                '-ios' in index[0].lower() or
                'ios-' in index[0].lower() or
                'ios/' in index[0].lower() or
                '/ios' in index[0].lower()):
                years[index[6].year] += 1
        else:
            years[index[6].year] += 1
    return [[year, count] for year, count in sorted(years.items())]


def top_skills_by_year(data):
    skills_frame = data[['year','key_skills']].dropna().copy()
    years_list = list(skills_frame['year'].unique())
    skills_by_year = {}

    for year in years_list:
        correct_skills = []
        year_skills = skills_frame[skills_frame['year'] == year]['key_skills'].str.cat(sep='\n')
        year_skills = re.split(r', |\n', year_skills)

        for skill in year_skills:
            skill = skill.strip()
            correct_skills.append(skill)
        skills_by_year[year] = correct_skills

    top_skills = {}
    for year, skills in skills_by_year.items():
        top_skill = {}
        for skill in skills:
            top_skill[skill] = top_skill.get(skill, 0) + 1
        sorted_skills = sorted(top_skill.items(), key=lambda x: x[1], reverse=True)[:20]
        top_skills[year] = sorted_skills

    return top_skills

def top_skills_by_year_all(data):
    skills_frame = data[['year','key_skills']].dropna().copy()
    years_list = list(skills_frame['year'].unique())
    skills_by_year = {}

    for year in years_list:
        correct_skills = []
        year_skills = skills_frame[skills_frame['year'] == year]['key_skills'].str.cat(sep='\n')
        year_skills = re.split(r', |\n', year_skills)

        for skill in year_skills:
            skill = skill.strip()
            correct_skills.append(skill)
        skills_by_year[year] = correct_skills

    top_skills_all = {}
    for year, skills in skills_by_year.items():
        top_skill = {}
        for skill in skills:
            top_skill[skill] = top_skill.get(skill, 0) + 1
        sorted_skills = sorted(top_skill.items(), key=lambda x: x[1], reverse=True)[:20]
        top_skills_all[year] = sorted_skills

    return top_skills_all


def create_report(data, profession, stats_salary_all, stats_salary_prof, stats_vacancies_all, stats_vacancies_prof, top_skills, city_salary_all, city_salary_prof, sorted_cities_share_all, sorted_cities_share_prof, top_skills_all):
    """Создает Excel-отчет."""
    workbook = Workbook()

    sheet1 = workbook.create_sheet(title='Динамика уровня зарплат по годам (All)')
    sheet1.append(['Год', 'Средняя зарплата', 'Вакансии'])
    for row in stats_salary_all:
        sheet1.append(row)

    sheet2 = workbook.create_sheet(title=f'Динамика уровня зарплат по годам ({profession})')
    sheet2.append(['Год', 'Средняя зарплата', 'Вакансии'])
    for row in stats_salary_prof:
        sheet2.append(row)

    sheet3 = workbook.create_sheet(title='Кол-во вакансий')
    sheet3.append(['Год', 'Вакансии'])
    for row in stats_vacancies_all:
        sheet3.append(row)

    sheet4 = workbook.create_sheet(title=f'Кол-во вакансий ({profession})')
    sheet4.append(['Год', 'Вакансии'])
    for row in stats_vacancies_prof:
        sheet4.append(row)

    sheet5 = workbook.create_sheet(title=f'Навыки ({profession})')
    sheet5.append(['Год', 'Навыки', 'Частота'])
    for year, skills in top_skills.items():
        for skill, freq in skills:
            sheet5.append([year, skill, freq])

    sheet6 = workbook.create_sheet(title='Зарплата по городам')
    sheet6.append(['Город', 'Средняя зарплата'])
    for row in city_salary_all:
        sheet6.append(row)

    sheet7 = workbook.create_sheet(title=f'Зарплата по городам ({profession})')
    sheet7.append(['Город', 'Средняя зарплата'])
    for row in city_salary_prof:
        sheet7.append(row)

    sheet8 = workbook.create_sheet(title='Вакансии по городам')
    sheet8.append(['Город', 'Вакансии'])
    for row in sorted_cities_share_all:
        sheet8.append(row)

    sheet9 = workbook.create_sheet(title=f'Вакансии по городам ({profession})')
    sheet9.append(['Город', 'Вакансии'])
    for row in sorted_cities_share_prof:
        sheet9.append(row)
        
    sheet10 = workbook.create_sheet(title=f'Навыки ({profession})')
    sheet10.append(['Год', 'Навыки', 'Частота'])
    for year, skills in top_skills_all.items():
        for skill, freq in skills:
            sheet10.append([year, skill, freq])

    workbook.save('analytics.xlsx')


def process_file(file_path, profession=None):
    data = pd.read_csv(file_path)
    data['published_at'] = pd.to_datetime(data['published_at'], errors='coerce', utc=True)
    data['year'] = data['published_at'].dt.year

    filtered_data = data[data['name'].str.lower().str.contains(profession) | data['name'].str.lower().str.contains('ios')]
    data_list = data[['name', 'key_skills', 'salary_from', 'salary_to', 'salary_currency', 'area_name', 'published_at', 'salary_rub']].values.tolist()

    stats_salary_all = prof_year_state_all(data_list)
    stats_salary_prof = prof_year_state_prof(data_list, profession)

    stats_vacancies_all = vacancy_count_by_year(data_list)
    stats_vacancies_prof = vacancy_count_by_year(data_list, profession)

    top_skills = top_skills_by_year(filtered_data)
    top_skills_all = top_skills_by_year_all(data)

    city_salary_all, city_salary_prof = plot_city_salary_dynamics(data_list, profession)
    sorted_cities_share_all, sorted_cities_share_prof = plot_city_vacancy_share(data_list, profession)

    create_report(data, profession, stats_salary_all, stats_salary_prof, stats_vacancies_all, stats_vacancies_prof, top_skills, city_salary_all, city_salary_prof, sorted_cities_share_all, sorted_cities_share_prof, top_skills_all)
    generate_all_plots(data_list, profession, stats_salary_all, stats_salary_prof, stats_vacancies_all, stats_vacancies_prof, top_skills, top_skills_all)


def plot_salary_dynamics(stats_salary_all, stats_salary_prof, profession):
    """Графики динамики уровня зарплат по годам."""
    years_all = [row[0] for row in stats_salary_all]
    salaries_all = [row[1] for row in stats_salary_all]
    years_prof = [row[0] for row in stats_salary_prof]
    salaries_prof = [row[1] for row in stats_salary_prof]

    plt.figure(figsize=(10, 6))
    plt.bar(years_all, salaries_all, color='#988EC8')
    plt.title('Динамика уровня зарплат по годам для всех вакансий')
    plt.ylabel('Средняя зарплата (RUB)')
    plt.xticks(years_all, fontsize=10, rotation=90)
    plt.yticks(fontsize=10)
    plt.grid(axis='y',linestyle='--',alpha=0.5)
    plt.savefig('salary_dynamics_all.png')
    
    plt.figure(figsize=(10, 6))
    plt.bar(years_prof, salaries_prof, color='#988EC8')
    plt.title(f'Динамика уровня зарплат по годам для {profession}')
    plt.ylabel('Средняя зарплата (RUB)')
    plt.xticks(years_prof, fontsize=10, rotation=90)
    plt.yticks(fontsize=10)
    plt.grid(axis='y',linestyle='--',alpha=0.5)
    plt.savefig('salary_dynamics_prof.png')


def plot_vacancy_dynamics(stats_vacancies_all, stats_vacancies_prof, profession):
    """Графики динамики количества вакансий по годам."""
    years_all = [row[0] for row in stats_vacancies_all]
    counts_all = [row[1] for row in stats_vacancies_all]
    years_prof = [row[0] for row in stats_vacancies_prof]
    counts_prof = [row[1] for row in stats_vacancies_prof]

    plt.figure(figsize=(10, 6))
    plt.bar(years_all, counts_all, color='#988EC8')
    plt.title('Динамика количества вакансий по годам для всех вакансий', fontsize=16)
    plt.ylabel('Количество вакансий', fontsize=14)
    plt.xticks(years_all, fontsize=10, rotation=90)
    plt.yticks(fontsize=10)
    plt.grid(axis='y',linestyle='--',alpha=0.5)
    plt.savefig('vacancy_dynamics_all.png')

    plt.figure(figsize=(10, 6))
    plt.bar(years_prof, counts_prof, color='#988EC8')
    plt.title(f'Динамика количества вакансий по годам для {profession}', fontsize=16)
    plt.ylabel('Количество вакансий', fontsize=14)
    plt.xticks(years_prof, fontsize=10, rotation=90)
    plt.yticks(fontsize=10)
    plt.grid(axis='y',linestyle='--',alpha=0.5)
    plt.savefig('vacancy_dynamics_prof.png')
    

def plot_city_salary_dynamics(data, profession=None):
    """График уровня зарплат по городам для всех профессий и выбранной профессии."""
    city_salaries_all = {}
    city_salaries_prof = {}

    for index in data:
        if len(index) <= 5 or not index[5]:
            continue
        
        city = index[5]
        salary_rub = index[7]

        if salary_rub and salary_rub < 10000000:
            if city_salaries_all.get(city):
                city_salaries_all[city].append(salary_rub)
            else:
                city_salaries_all[city] = [salary_rub]
            if profession and (
                ' ios' in index[0].lower() or
                'ios ' in index[0].lower() or
                '-ios' in index[0].lower() or
                'ios-' in index[0].lower() or
                'ios/' in index[0].lower() or
                '/ios' in index[0].lower()):
                if city_salaries_prof.get(city):
                    city_salaries_prof[city].append(salary_rub)
                else:
                    city_salaries_prof[city] = [salary_rub]

    average_salaries_all = {
        city: sum(salaries) / len(salaries)
        for city, salaries in city_salaries_all.items()
    }
    average_salaries_prof = {
        city: sum(salaries) / len(salaries)
        for city, salaries in city_salaries_prof.items()
    }

    sorted_cities_all = sorted(average_salaries_all.items(), key=lambda x: x[1], reverse=True)[:20]
    sorted_cities_prof = sorted(average_salaries_prof.items(), key=lambda x: x[1], reverse=True)[:20]

    cities_all = [item[0] for item in sorted_cities_all]
    salaries_all = [item[1] for item in sorted_cities_all]
    cities_prof = [item[0] for item in sorted_cities_prof]
    salaries_prof = [item[1] for item in sorted_cities_prof]

    plt.figure(figsize=(10, 6))
    plt.barh(cities_all, salaries_all, color='#988EC8')
    plt.title(f'Уровень зарплат по городам для всех вакансий', fontsize=16)
    plt.xlabel('Средняя зарплата (RUB)', fontsize=14)
    plt.grid(axis='x',linestyle='--',alpha=0.5)
    plt.gca().invert_yaxis()
    plt.tight_layout() 
    plt.savefig('city_salary_dynamics_all.png')

    plt.figure(figsize=(10, 6))
    plt.barh(cities_prof, salaries_prof, color='#988EC8')
    plt.title(f'Уровень зарплат по городам для {profession}', fontsize=16)
    plt.xlabel('Средняя зарплата (RUB)', fontsize=14)
    plt.grid(axis='x',linestyle='--',alpha=0.5)
    plt.gca().invert_yaxis()
    plt.tight_layout()  
    plt.savefig('city_salary_dynamics_prof.png')

    return sorted_cities_all, sorted_cities_prof
    

def plot_city_vacancy_share(data, profession=None):
    """График доли вакансий по городам, включая категории с долей < 1%."""
    city_vacancies_all = {}
    city_vacancies_prof = {}
    total_vacancies_all = 0
    total_vacancies_prof = 0

    for index in data:
        if len(index) > 5 and index[5]:
            city = index[5]
            if city not in city_vacancies_all:
                city_vacancies_all[city] = 1
            else:
                city_vacancies_all[city] += 1
            total_vacancies_all += 1

            if profession and (
                ' ios' in index[0].lower() or
                'ios ' in index[0].lower() or
                '-ios' in index[0].lower() or
                'ios-' in index[0].lower() or
                'ios/' in index[0].lower() or
                '/ios' in index[0].lower()):
                if city not in city_vacancies_prof:
                    city_vacancies_prof[city] = 1
                else:
                    city_vacancies_prof[city] += 1
                total_vacancies_prof += 1

    city_shares_all = {
        city: (count / total_vacancies_all) * 100
        for city, count in city_vacancies_all.items()
    }

    city_shares_prof = {}
    if total_vacancies_prof > 0:
        city_shares_prof = {
            city: (count / total_vacancies_prof) * 100
            for city, count in city_vacancies_prof.items()
        }

    def categorize_cities(shares_dict):
        sorted_shares = sorted(shares_dict.items(), key=lambda x: x[1], reverse=True)
        other_cities_share = sum(share for city, share in sorted_shares if share < 1)
        filtered_shares = [(city, share) for city, share in sorted_shares if share >= 1]

        if other_cities_share > 0:
            filtered_shares.append(("Другие", other_cities_share))

        return filtered_shares

    sorted_cities_share_all = categorize_cities(city_shares_all)
    sorted_cities_share_prof = categorize_cities(city_shares_prof)

    if sorted_cities_share_all:
        cities_all = [item[0] for item in sorted_cities_share_all]
        shares_all = [item[1] for item in sorted_cities_share_all]
        plt.figure(figsize=(10, 6))
        plt.pie(shares_all, labels=cities_all, autopct='%1.1f%%', startangle=90)
        plt.title('Доля вакансий по городам для всех профессий')
        plt.savefig('city_vacancy_share_all.png')
        plt.close()

    if sorted_cities_share_prof:
        cities_prof = [item[0] for item in sorted_cities_share_prof]
        shares_prof = [item[1] for item in sorted_cities_share_prof]
        plt.figure(figsize=(10, 6))
        plt.pie(shares_prof, labels=cities_prof, autopct='%1.1f%%', startangle=90)
        plt.title(f'Доля вакансий по городам для {profession}')
        plt.savefig('city_vacancy_share_prof.png')
        plt.close()

    return sorted_cities_share_all, sorted_cities_share_prof



def plot_top_skills(top_skills, top_skills_all, profession):
    """График ТОП-20 навыков по годам для выбранной профессии."""
    for year, skills in top_skills.items():
        skill_names = [item[0] for item in skills]
        frequencies = [item[1] for item in skills]

        plt.figure(figsize=(12, 8))
        plt.barh(skill_names, frequencies, color='#988EC8')
        plt.title(f'Топ 20 навыков в {year} для {profession}', fontsize=16)
        plt.xlabel('Частотность', fontsize=14)
        plt.grid(axis='x',linestyle='--',alpha=0.5)
        plt.gca().invert_yaxis()
        plt.tight_layout()
        plt.savefig(f'top_skills_{year}.png')

    for year, skills in top_skills_all.items():
        skill_names = [item[0] for item in skills]
        frequencies = [item[1] for item in skills]

        plt.figure(figsize=(12, 8))
        plt.barh(skill_names, frequencies, color='#988EC8')
        plt.title(f'Топ 20 навыков в {year} для всех профессий', fontsize=16)
        plt.xlabel('Частотность', fontsize=14)
        plt.grid(axis='x',linestyle='--',alpha=0.5)
        plt.gca().invert_yaxis()
        plt.tight_layout()
        plt.savefig(f'top_skills_all_{year}.png')
        

def generate_all_plots(filtered_data, profession, stats_salary_all, stats_salary_prof, stats_vacancies_all, stats_vacancies_prof, top_skills, top_skills_all):
    plot_salary_dynamics(stats_salary_all, stats_salary_prof, profession)
    plot_vacancy_dynamics(stats_vacancies_all, stats_vacancies_prof, profession)
    plot_top_skills(top_skills, top_skills_all, profession)


process_file('vacancies_in_rub.csv', profession='ios-разработчик')